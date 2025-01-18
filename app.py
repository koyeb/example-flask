cat << 'EOF' > app.py
import os
import logging
import tempfile
from flask import Flask, redirect, url_for, send_file, jsonify, render_template_string
from dotenv import load_dotenv
from office365.sharepoint.client_context import ClientContext
from openpyxl import load_workbook
import excel2img  # Assuming you are using this for generating images

app = Flask(__name__)

# Load environment variables
load_dotenv()

# SharePoint credentials and file paths
SHAREPOINT_SITE_URL = os.getenv('SHAREPOINT_SITE_URL')
SHAREPOINT_FILE_URL = os.getenv('SHAREPOINT_FILE_URL')
SHAREPOINT_USERNAME = os.getenv('SHAREPOINT_USERNAME')
SHAREPOINT_PASSWORD = os.getenv('SHAREPOINT_PASSWORD')

# Logger setup
logger = logging.getLogger('FlaskAppLogger')
logger.setLevel(logging.INFO)

console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)

# Log to a temporary directory
with tempfile.TemporaryDirectory() as temp_image_folder:
    error_log_path = os.path.join(temp_image_folder, 'Error.log')
    file_handler = logging.FileHandler(error_log_path)
    file_handler.setLevel(logging.ERROR)

    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    console_handler.setFormatter(formatter)
    file_handler.setFormatter(formatter)

    logger.addHandler(console_handler)
    logger.addHandler(file_handler)

    logger.info(f"Temporary directory for logs created at {temp_image_folder}")


def download_sharepoint_file():
    """Download the Excel file from SharePoint"""
    try:
        # Set up SharePoint authentication
        ctx = ClientContext(SHAREPOINT_SITE_URL).with_credentials(SHAREPOINT_USERNAME, SHAREPOINT_PASSWORD)
        
        # Fetch the file from SharePoint
        file = ctx.web.get_file_by_server_relative_url(SHAREPOINT_FILE_URL)
        ctx.load(file)
        ctx.execute_query()

        # Use tempfile to create a temporary file for the downloaded Excel file
        with tempfile.NamedTemporaryFile(suffix='.xlsm', delete=False) as temp_file:
            temp_file_path = temp_file.name
            file.download(temp_file)
            ctx.execute_query()
        
        logger.info(f"File downloaded successfully to {temp_file_path}")
        return temp_file_path
    except Exception as e:
        logger.error(f"Failed to download SharePoint file: {e}")
        return None


def load_workbook_simple(file_path):
    """Load an unprotected Excel workbook using openpyxl"""
    try:
        workbook = load_workbook(file_path, data_only=True)
        logger.info(f"Successfully loaded workbook: {file_path}")
        return workbook
    except Exception as e:
        logger.error(f"Failed to load workbook '{file_path}': {e}")
        return None


def generate_image(sheet_name, cell_range, image_path):
    """Generate an image from an Excel sheet"""
    try:
        # Download the file from SharePoint
        excel_file_path = download_sharepoint_file()
        if not excel_file_path:
            raise Exception("Unable to download Excel file from SharePoint.")
        
        # Load workbook
        workbook = load_workbook_simple(excel_file_path)
        if not workbook:
            raise Exception("Unable to load workbook for image export.")
        
        # Check if the sheet exists
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
        
        # Hide gridlines
        wb_sheet = workbook[sheet_name]
        wb_sheet.sheet_view.showGridLines = False

        # Save the workbook to a temporary file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            temp_decrypted_path = tmp.name
            workbook.save(temp_decrypted_path)

        # Generate image using excel2img
        excel2img.export_img(temp_decrypted_path, image_path, sheet_name, cell_range)
        logger.info(f"Generated image for '{sheet_name}' -> {image_path}")

    except Exception as e:
        logger.error(f"Error generating image for '{sheet_name}': {e}")
        raise e


@app.route('/')
def home():
    """Main route to render the web page"""
    # Determine which pages are active (this will be adjusted according to your rules)
    active_pages = ['Morning', 'Evening', 'Night', 'Friday']

    # Sheet ranges for active pages
    page_ranges = {
        'Morning': ('Morning', 'A1:H33'),
        'Evening': ('Evening', 'A1:H33'),
        'Night': ('Night', 'A1:H33'),
        'Friday': ('Friday', 'A1:H33')
    }

    pages = []
    for page in active_pages:
        if page not in page_ranges:
            logger.error(f"No cell range defined for page '{page}'. Skipping.")
            continue

        sheet_name, cell_range = page_ranges[page]
        image_filename = f"{page}.png"
        
        # Use tempfile to create a temporary image folder dynamically
        with tempfile.TemporaryDirectory() as temp_image_folder:
            image_path = os.path.join(temp_image_folder, image_filename)

            try:
                generate_image(sheet_name, cell_range, image_path)
                image_url = url_for('serve_image', filename=image_filename)
                pages.append({'url': image_url, 'header': f"{sheet_name} Shift"})
            except Exception as e:
                logger.error(f"Failed to generate image for page '{page}': {e}")

    if not pages:
        return "<h1>No images available for the active pages.</h1>"

    # HTML content for the page
    html_content = f"""
    <html>
    <head>
      <style>
        html, body {{
          margin: 0; padding: 0;
          width: 100%; height: 100%;
          background-color: white;
          display: flex; flex-direction: column; align-items: center;
        }}
        #header {{
          position: fixed;
          top: 10px;
          left: 0; right: 0;
          text-align: center;
          font-size: 24px;
          color: black;
          padding: 10px 0;
        }}
        #image-container {{
          width: 90vw;
          height: 90vh;
          margin-top: 60px;
          display: flex; justify-content: center; align-items: center;
        }}
        img {{
          max-width: 100%;
          max-height: 100%;
          object-fit: contain;
          display: none;
        }}
        img.active {{
          display: block;
        }}
      </style>
    </head>
    <body>
      <div id="header">Loading...</div>
      <div id="image-container"></div>
      <script>
        const pages = {pages};
        let currentIndex = 0;
        
        window.onload = function() {{
          const header = document.getElementById('header');
          const imageContainer = document.getElementById('image-container');

          const imgElements = pages.map(page => {{
            const img = document.createElement('img');
            img.src = page.url;
            imageContainer.appendChild(img);
            return img;
          }}); 

          function showPage(idx) {{
            imgElements.forEach(img => img.classList.remove('active'));
            imgElements[idx].classList.add('active');
            header.innerText = pages[idx].header;
          }}

          setInterval(() => {{
            currentIndex = (currentIndex + 1) % pages.length;
            showPage(currentIndex);
          }}, 30000);

          showPage(currentIndex);
        }};</script>
    </body>
    </html>
    """

    return render_template_string(html_content, pages=pages)


@app.route('/temp_images/<filename>')
def serve_image(filename):
    """Serve generated images"""
    try:
        # If you have a dedicated temp folder or want an in-memory approach,
        # adjust accordingly. For demonstration, using non-existent variable:
        return send_file(os.path.join(TEMP_IMAGE_FOLDER, filename))
    except Exception as e:
        logger.error(f"Failed to serve image {filename}: {e}")
        return "Error serving image.", 500


if __name__ == '__main__':
    app.run(debug=True)
EOF
